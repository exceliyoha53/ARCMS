import io
import logging
from openpyxl import load_workbook
from app.auth import require_registrar, hash_password
from app.database import get_connection, return_connection, get_db_cursor
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/import", tags=["Data Import"])


def parse_students_excel(file_buffer: io.BytesIO) -> tuple[list[dict], list[dict]]:
    """
    Parses a student bulk import Excel file.
    Expected columns: matric_number, first_name, last_name, email,
                      department_code, entry_year, level

    Parameters:
        file_buffer (BytesIO): In-memory Excel file

    Returns:
        tuple[list[dict], list[dict]]: (valid_rows, errors)
    """

    valid_rows = []
    errors = []

    try:
        wb = load_workbook(file_buffer, data_only=True, read_only=True)
        ws = wb.active

    except Exception as e:
        return [], [{"row": 0, "error": f"Cannot read file: {e}"}]

    headers = [
        str(cell).strip().lower() if cell else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    required = {
        "matric_number",
        "first_name",
        "last_name",
        "email",
        "department_code",
        "entry_year",
        "level",
    }

    missing = required - set(headers)
    if missing:
        return [], [{"row": 1, "error": f"Missing columns: {', '.join(missing)}"}]
    # this way the code wont break if headers are not arranged
    col = {name: headers.index(name) for name in required}

    # start reading at row 2, track the current row number
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue  # skip empty rows

        try:
            record = {
                "matric_number": str(row[col["matric_number"]]).strip().upper(),
                "first_name": str(row[col["first_name"]]).strip(),
                "last_name": str(row[col["last_name"]]).strip(),
                "email": str(row[col["email"]]).strip().lower(),
                "department_code": str(row[col["department_code"]]).strip().upper(),
                "entry_year": int(row[col["entry_year"]]),
                "level": int(row[col["level"]]),
            }

            if record["level"] not in {100, 200, 300, 400, 500}:
                raise ValueError(f"Invalid level: {record['level']}")
            valid_rows.append(record)

        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})
    return valid_rows, errors


def parse_historical_scores_excel(
    file_buffer: io.BytesIO,
) -> tuple[list[dict], list[dict]]:
    """
    Parses a historical scores Excel file for backdating results.
    Expected columns: matric_number, course_code, session_name,
                      semester_number, ca_score, exam_score, level

    Parameters:
        file_buffer (BytesIO): In-memory Excel file

    Returns:
        tuple[list[dict], list[dict]]: (valid_rows, errors)
    """
    valid_rows = []
    errors = []

    try:
        wb = load_workbook(file_buffer, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return [], [{"row": 0, "error": f"Cannot read file: {e}"}]

    headers = [
        str(cell).strip().lower() if cell else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    required = {
        "matric_number",
        "course_code",
        "session_name",
        "semester_number",
        "ca_score",
        "exam_score",
        "level",
    }
    missing = required - set(headers)
    if missing:
        return [], [{"row": 1, "error": f"Missing columns: {', '.join(missing)}"}]

    col = {name: headers.index(name) for name in required}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        try:
            ca = float(row[col["ca_score"]])
            exam = float(row[col["exam_score"]])
            if not (0 <= ca <= 30):
                raise ValueError(f"CA score {ca} out of range")
            if not (0 <= exam <= 70):
                raise ValueError(f"Exam score {exam} out of range")

            valid_rows.append(
                {
                    "matric_number": str(row[col["matric_number"]]).strip().upper(),
                    "course_code": str(row[col["course_code"]]).strip().upper(),
                    "session_name": str(row[col["session_name"]]).strip(),
                    "semester_number": int(row[col["semester_number"]]),
                    "ca_score": ca,
                    "exam_score": exam,
                    "level": int(row[col["level"]]),
                }
            )

        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    return valid_rows, errors


@router.post("/students", status_code=status.HTTP_201_CREATED)
def bulk_import_students(
    file: UploadFile = File(...), current_user: dict = Depends(require_registrar)
) -> dict:
    """
    Bulk imports students from an Excel file.
    Creates a user account and student profile for each row.
    Default password is the matric number — student must change on first login.
    Skips rows where matric number or email already exists.

    Expected columns: matric_number, first_name, last_name, email,
                      department_code, entry_year, level

    Parameters:
        file (UploadFile): Excel file (.xlsx)
        current_user (dict): Injected by require_registrar

    Returns:
        dict: Summary with created_count, skipped_count, errors
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files accepted"
        )

    file_buffer = io.BytesIO(file.file.read())
    valid_rows, errors = parse_students_excel(file_buffer)

    if not valid_rows and errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "File parsing failed", "errors": errors},
        )

    conn = get_connection()
    cursor = get_db_cursor(conn)
    created = 0
    skipped = 0
    row_errors = list(errors)

    try:
        for row in valid_rows:
            try:
                # get department ID from code
                cursor.execute(
                    "SELECT id FROM departments WHERE code = %s",
                    (row["department_code"],),
                )
                dept = cursor.fetchone()
                if not dept:
                    row_errors.append(
                        {
                            "matric": row["matric_number"],
                            "error": f"Department code {row['department_code']} not found",
                        }
                    )
                    skipped += 1
                    continue

                cursor.execute(
                    "SELECT id FROM pprogrammes WHERE department_id = %s LIMIT 1",
                    (dept["id"],),
                )
                prog = cursor.fetchone()
                if not prog:
                    row_errors.append(
                        {
                            "matric": row["matric_number"],
                            "error": "No programme found for department",
                        }
                    )
                    skipped += 1
                    continue
                # skip if matric already exists
                cursor.execute(
                    "SELECT id FROM students WHERE matric_number = %s",
                    (row["matric_number"],),
                )
                if cursor.fetchone():
                    skipped += 1
                    continue

                # create user account — default password is matric number
                hashed = hash_password(row["matric_number"])
                cursor.execute(
                    """
                    INSERT INTO users (email, password_hash, role)
                    VALUES (%s, %s, 'student')
                    RETURNING id
                """,
                    (row["email"], hashed),
                )
                user_id = cursor.fetchone()["id"]

                # create student profile
                cursor.execute(
                    """
                    INSERT INTO students (
                        user_id, matric_number, first_name, last_name,
                        department_id, programme_id, current_level, entry_year
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                    (
                        user_id,
                        row["matric_number"],
                        row["first_name"],
                        row["last_name"],
                        dept["id"],
                        prog["id"],
                        row["level"],
                        row["entry_year"],
                    ),
                )

                created += 1

            except Exception as e:
                row_errors.append(
                    {"matric": row.get("matric_number", "unknown"), "error": str(e)}
                )
                skipped += 1

        conn.commit()
        logger.info(
            f"Bulk import by {current_user['email']}: "
            f"{created} created, {skipped} skipped"
        )

        return {
            "message": "Import complete",
            "created": created,
            "skipped": skipped,
            "errors": row_errors,
        }

    except Exception as e:
        conn.rollback()
        logger.error(f"Bulk import error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Import failed"
        )

    finally:
        cursor.close()
        return_connection(conn)


@router.post("historical-scores", status_code=status.HTTP_201_CREATED)
def import_historical_scores(
    file: UploadFile = File(...), current_user: dict = Depends(require_registrar)
) -> dict:
    """
    Imports historical scores for existing students.
    Used to backdate results for students who were enrolled before ARCMS.
    Creates sessions, semesters, course offerings, registrations,
    and score records as needed. All imported scores are auto-approved.

    Expected columns: matric_number, course_code, session_name,
                      semester_number, ca_score, exam_score, level

    Parameters:
        file (UploadFile): Excel file (.xlsx)
        current_user (dict): Injected by require_registrar

    Returns:
        dict: Summary with imported_count and errors
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files accepted"
        )

    file_buffer = io.BytesIO(file.file.read())
    valid_rows, errors = parse_historical_scores_excel(file_buffer)

    if not valid_rows and errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "File parsing failed", "errors": errors},
        )

    conn = get_connection()
    cursor = get_db_cursor(conn)
    imported = 0
    skipped = 0
    row_errors = list(errors)

    try:
        from app.gpa_engine import get_grade

        for row in valid_rows:
            try:
                # get student
                cursor.execute(
                    "SELECT id FROM students WHERE matric_number = %s",
                    (row["matric_number"],),
                )
                student = cursor.fetchone()
                if not student:
                    row_errors.append(
                        {
                            "matric": row["matric_number"],
                            "error": "Student not found in system",
                        }
                    )
                    skipped += 1
                    continue

                # get or create session
                cursor.execute(
                    "SELECT id FROM sessions WHERE name = %s", (row["session_name"],)
                )
                session = cursor.fetchone()
                if not session:
                    cursor.execute(
                        """
                        INSERT INTO sessions (name, is_current)
                        VALUES (%s, FALSE) RETURNING id
                    """,
                        (row["session_name"],),
                    )
                    session = cursor.fetchone()

                # get or create semester
                cursor.execute(
                    """
                    SELECT id FROM semesters
                    WHERE session_id = %s AND semester_number = %s
                """,
                    (session["id"], row["semester_number"]),
                )
                semester = cursor.fetchone()
                if not semester:
                    cursor.execute(
                        """
                        INSERT INTO semesters (session_id, semester_number, is_current)
                        VALUES (%s, %s, FALSE) RETURNING id
                    """,
                        (session["id"], row["semester_number"]),
                    )
                    semester = cursor.fetchone()

                # get course
                cursor.execute(
                    "SELECT id FROM courses WHERE code = %s", (row["course_code"],)
                )
                course = cursor.fetchone()
                if not course:
                    row_errors.append(
                        {
                            "matric": row["matric_number"],
                            "error": f"Course {row['course_code']} not found",
                        }
                    )
                    skipped += 1
                    continue

                # get or create course offering
                cursor.execute(
                    """
                    SELECT id FROM course_offerings
                    WHERE course_id = %s AND semester_id = %s
                """,
                    (course["id"], semester["id"]),
                )
                offering = cursor.fetchone()
                if not offering:
                    cursor.execute(
                        """
                        INSERT INTO course_offerings (course_id, semester_id, lecturer_id)
                        VALUES (%s, %s, NULL) RETURNING id
                    """,
                        (course["id"], semester["id"]),
                    )
                    offering = cursor.fetchone()

                # get or create registration
                cursor.execute(
                    """
                    SELECT id FROM registrations
                    WHERE student_id = %s AND course_offering_id = %s
                """,
                    (student["id"], offering["id"]),
                )
                registration = cursor.fetchone()
                if not registration:
                    cursor.execute(
                        """
                        INSERT INTO registrations
                        (student_id, course_offering_id, level)
                        VALUES (%s, %s, %s) RETURNING id
                    """,
                        (student["id"], offering["id"], row["level"]),
                    )
                    registration = cursor.fetchone()

                # calculate grade and insert score — auto-approved
                total = row["ca_score"] + row["exam_score"]
                grade, grade_point = get_grade(total)

                cursor.execute(
                    """
                    INSERT INTO scores (
                        registration_id, ca_score, exam_score,
                        grade, grade_point, approval_status
                    )
                    VALUES (%s, %s, %s, %s, %s, 'approved')
                    ON CONFLICT (registration_id) DO UPDATE SET
                        ca_score = EXCLUDED.ca_score,
                        exam_score = EXCLUDED.exam_score,
                        grade = EXCLUDED.grade,
                        grade_point = EXCLUDED.grade_point,
                        approval_status = 'approved'
                """,
                    (
                        registration["id"],
                        row["ca_score"],
                        row["exam_score"],
                        grade,
                        grade_point,
                    ),
                )

                imported += 1

            except Exception as e:
                row_errors.append(
                    {"matric": row.get("matric_number", "unknown"), "error": str(e)}
                )
                skipped += 1

        conn.commit()

        affected_matrics = list(set(r["matric_number"] for r in valid_rows))
        from app.services.gpa_service import recalculate_student_gpa

        for matric in affected_matrics:
            cursor2 = get_db_cursor(conn)
            cursor2.execute(
                "SELECT id FROM students WHERE matric_number = %s", (matric,)
            )
            s = cursor2.fetchone()
            cursor2.close()
            if s:
                recalculate_student_gpa(s["id"], conn)
        conn.commit()

        logger.info(
            f"Historical import by {current_user['email']}: "
            f"{imported} imported, {skipped} skipped"
        )

        return {
            "message": "Historical import complete",
            "imported": imported,
            "skipped": skipped,
            "errors": row_errors,
        }
    except Exception as e:
        conn.rollback()
        logger.error(f"Historical import error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Historical import failed",
        )

    finally:
        cursor.close()
        return_connection(conn)
