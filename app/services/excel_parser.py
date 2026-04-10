import logging
from openpyxl import load_workbook  # reads .xlsx files
from io import BytesIO
from app.models.schemas import ScoreUpload
from pydantic import ValidationError

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {
    "matric_number",
    "ca_score",
    "exam_score",
}  # column headers case insensitive


def parse_score_sheet(file_buffer: BytesIO) -> tuple[list[ScoreUpload], list[dict]]:
    """
    Reads an Excel score sheet and validates every row.
    Returns validated score objects and a list of row-level errors.
    Rejects nothing silently — every problem is reported back to the lecturer.

    The Excel file must have these exact column headers (case-insensitive):
        matric_number — student's AFIT matric number
        ca_score      — continuous assessment score (0-30)
        exam_score    — examination score (0-70)

    Parameters:
        file_buffer (BytesIO): In-memory Excel file bytes from the upload

    Returns:
        tuple[list[ScoreUpload], list[dict]]:
            - list of validated ScoreUpload objects ready for database insertion
            - list of error dicts with keys: row, matric_number, error
              empty list means all rows passed validation
    """
    parsed_scores = []
    errors = []

    try:
        workbook = load_workbook(
            file_buffer, data_only=True, read_only=True
        )  # True here returns cell values not formulas
        sheet = workbook.active  # use first sheet

    except Exception as e:
        logger.error(f"failed to open excel file: {e}")

        return [], [
            {"row": 0, "matric_number": "N/A", "error": f"Could not read file: {e}"}
        ]  # return one top level error

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in sheet[1]  # sheet[1] is first row
    ]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return [], [
            {
                "row": 1,
                "matric_number": "N/A",
                "error": f"Missing required columns: {', '.join(missing)}",
            }
        ]

    # mapping the column names to index positions
    # looks like {"matric_number": 0, "ca_score": 1, "exam_score": 2}
    col_index = {name: headers.index(name) for name in REQUIRED_COLUMNS}

    # process data rows - sheet starts at row 1, data starts at row 2
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(cell is None for cell in row):
            continue  # skips empty rows

        matric_raw = row[col_index["matric_number"]]
        ca_raw = row[col_index["ca_score"]]
        exam_raw = row[col_index["exam_score"]]

        # convert matric number to string
        matric = str(matric_raw).strip() if matric_raw is not None else ""

        try:
            ca = float(ca_raw) if ca_raw is not None else None
            exam = float(exam_raw) if exam_raw is not None else None

            if ca is None or exam is None:
                raise ValueError("CA score or exam score is missing")

            score = ScoreUpload(matric_number=matric, ca_score=ca, exam_score=exam)
            parsed_scores.append(score)

        except (
            ValidationError
        ) as e:  # pydantic validation failed, extract readable error msg
            error_messages = "; ".join(err["msg"] for err in e.errors())

            errors.append(
                {
                    "row": row_num,
                    "matric_number": matric or "unknown",
                    "error": error_messages,
                }
            )
            logger.warning(f"Row {row_num} validation failed: {error_messages}")

        except ValueError as e:
            errors.append(
                {"row": row_num, "matric_number": matric or "unknown", "error": str(e)}
            )
            logger.warning(f"Row {row_num} value error: {e}")

    logger.info(
        f"Excel parse complete - {len(parsed_scores)} valid rows, {len(errors)} errors"
    )
    return parsed_scores, errors
